"""
run_all_flant5.py

Runs Magnitude / Movement / SparseGPT / HAWQ / ZeroQuant / GAMC compression
on Flan-T5 {small,base,large,xl} against the Yelp Review Full dataset,
and writes results in the same per-method block format as sheet4_yelp.

GAMC = the 4-stage pipeline (Teacher -> Pruned -> Distilled -> Quantized)
from run_all_flant5's original single-model script, generalized here
across all Flan-T5 sizes and folded into the same harness as the other
5 methods. Stage 1 (Teacher) reuses get_teacher_checkpoint, same as
every other method already does.

Does not modify flanutils.py. Yelp-specific data/eval logic lives in
yelp_utils.py.

Usage:
    python3 run_all_flant5.py
    python3 run_all_flant5.py --models flan-t5-small --methods GAMC Magnitude

CHANGES vs previous version:
  1. Per-model max_in/max_out from MODEL_CONFIGS are now actually passed
     into AGNewsSeq2SeqDataset (previously ignored -> everything ran at
     the dataset-default 256/16 regardless of model size, defeating the
     point of tuning those configs for Large/xl memory headroom).
  2. HAWQ now quantizes the fine-tuned teacher checkpoint instead of the
     raw pretrained model, matching every other method's behavior.
     bitsandbytes needs a full HF-format checkpoint (config.json +
     weights), not a bare state_dict, so get_teacher_checkpoint now also
     exports one via save_pretrained the first time it trains (or lazily,
     if reusing an old .pt-only checkpoint that predates this change).
  3. METHODS order matches the sheet4_yelp tab block order (SparseGPT,
     HAWQ, ZeroQuant, Magnitude, Movement, GAMC) so run output can be
     pasted in without reshuffling.
"""
import os
import gc
import argparse
import traceback

import torch
import torch.nn as nn
import pandas as pd
from transformers import AutoTokenizer, AutoModelForSeq2SeqLM
from torch.utils.data import DataLoader

from flanutils import (
    AGNewsSeq2SeqDataset,  # generic prompt/target tokenizer, name predates Yelp support
    task_weighted_prune, contrastive_distillation_loss, clear_gpu_memory,
)
from yelp_utils import load_yelp_data, prepare_dataset_yelp, evaluate_seq2seq_yelp

MODEL_CONFIGS = {
    "flan-t5-small": {
        "hf_name": "google/flan-t5-small",
        "batch_size": 32, "grad_accum": 1,
        "max_in": 256, "max_out": 16,
        "calib": 256, "train_n": 20000, "recovery_n": 3000,
        "test_batch": 64,
        "finetune_epochs": 3, "recovery_epochs": 2,
    },
    "flan-t5-base": {
        "hf_name": "google/flan-t5-base",
        "batch_size": 16, "grad_accum": 1,
        "max_in": 192, "max_out": 16,
        "calib": 192, "train_n": 8000, "recovery_n": 1500,
        "test_batch": 32,
        "finetune_epochs": 3, "recovery_epochs": 2,
    },
    "flan-t5-large": {
        "hf_name": "google/flan-t5-large",
        "batch_size": 8, "grad_accum": 2,
        "max_in": 128, "max_out": 8,
        "calib": 128, "train_n": 4000, "recovery_n": 1000,
        "test_batch": 16,
        "finetune_epochs": 2, "recovery_epochs": 1,
    },
    "flan-t5-xl": {
        "hf_name": "google/flan-t5-xl",
        "batch_size": 4, "grad_accum": 4,
        "max_in": 128, "max_out": 8,
        "calib": 96, "train_n": 1000, "recovery_n": 400,
        "test_batch": 8,
        # Left at 1 epoch each -- xl is already the highest OOM/runtime risk
        # in this harness (see GAMC Stage 3 checkpoint note below); bumping
        # its budget the same as small/base would multiply that risk instead
        # of just improving accuracy. Increase manually if you've confirmed
        # xl has comfortable headroom on this GPU.
        "finetune_epochs": 1, "recovery_epochs": 1,
    },
}
DISPLAY_NAMES = {
    "flan-t5-small": "Flan-T5-Small", "flan-t5-base": "Flan-T5-Base",
    "flan-t5-large": "Flan-T5-Large", "flan-t5-xl": "Flan-T5-xl",
}
# Order matches the sheet4_yelp tab's block order.
METHODS = ["SparseGPT", "HAWQ", "ZeroQuant", "Magnitude", "Movement", "GAMC"]

PRUNE_RATIO     = 0.20
LR              = 5e-5
CLIP_NORM       = 1.0
DISTILL_LR      = 1e-4


def clear_gpu():
    gc.collect()
    if torch.cuda.is_available():
        torch.cuda.empty_cache()


def get_prunable_params(model):
    return [(n, p) for n, p in model.named_parameters()
            if "weight" in n and p.dim() > 1 and "lm_head" not in n]


def make_loader(df, tokenizer, cfg, batch_size, shuffle=False):
    """Builds a DataLoader using this model's configured max_in/max_out
    instead of the AGNewsSeq2SeqDataset defaults."""
    ds = AGNewsSeq2SeqDataset(
        df, tokenizer,
        max_length=cfg["max_in"],
        target_max_length=cfg["max_out"],
    )
    return DataLoader(ds, batch_size=batch_size, shuffle=shuffle)


def _export_hf_dir(hf_name, teacher_path, teacher_hf_dir):
    """Materializes a full HF-format checkpoint (config + weights) from a
    raw state_dict so bitsandbytes' from_pretrained(quantization_config=...)
    can load the *fine-tuned* weights for HAWQ instead of the pretrained
    base model. from_pretrained is required here -- bnb can't quantize an
    already-instantiated model from a bare state_dict."""
    print(f"  [HAWQ] exporting HF-format checkpoint -> {teacher_hf_dir}")
    m = AutoModelForSeq2SeqLM.from_pretrained(hf_name, torch_dtype=torch.bfloat16)
    m.load_state_dict(torch.load(teacher_path, map_location="cpu"))
    m.save_pretrained(teacher_hf_dir)
    del m
    clear_gpu()


def get_teacher_checkpoint(model_key, hf_name, train_loader, device, grad_accum, finetune_epochs, force=False):
    artifacts_dir = f"artifacts-{model_key}-yelp"
    os.makedirs(artifacts_dir, exist_ok=True)
    teacher_path = os.path.join(artifacts_dir, "yelp_teacher.pt")
    teacher_hf_dir = os.path.join(artifacts_dir, "yelp_teacher_hf")

    if os.path.exists(teacher_path) and not force:
        print(f"  Found existing checkpoint for {model_key} - reusing")
        if not os.path.isdir(teacher_hf_dir):
            _export_hf_dir(hf_name, teacher_path, teacher_hf_dir)
        return teacher_path, teacher_hf_dir

    if force and os.path.exists(teacher_path):
        print(f"  --force_retrain_teacher set: retraining {model_key} from scratch "
              f"(finetune_epochs={finetune_epochs}), overwriting old checkpoint")
    else:
        print(f"  No checkpoint for {model_key} - training from scratch "
              f"(finetune_epochs={finetune_epochs})")
    model = AutoModelForSeq2SeqLM.from_pretrained(
        hf_name, torch_dtype=torch.bfloat16,
        device_map="cuda" if torch.cuda.is_available() else "auto",
    )
    optimizer = torch.optim.AdamW(model.parameters(), lr=LR)

    for epoch in range(finetune_epochs):
        model.train()
        if hasattr(model, "gradient_checkpointing_enable"):
            model.gradient_checkpointing_enable()
        optimizer.zero_grad()
        for i, batch in enumerate(train_loader):
            ids  = batch["input_ids"].to(device)
            mask = batch["attention_mask"].to(device)
            lbls = batch["labels"].to(device)
            with torch.amp.autocast(device_type="cuda", dtype=torch.bfloat16):
                loss = model(input_ids=ids, attention_mask=mask, labels=lbls).loss / grad_accum
            loss.backward()
            if (i + 1) % grad_accum == 0:
                torch.nn.utils.clip_grad_norm_(model.parameters(), CLIP_NORM)
                optimizer.step()
                optimizer.zero_grad()
            if i % 500 == 0:
                print(f"    [{model_key} ft] epoch {epoch} batch {i} loss {loss.item()*grad_accum:.4f}")

    torch.save(model.state_dict(), teacher_path)
    model.save_pretrained(teacher_hf_dir)
    del model
    clear_gpu()
    return teacher_path, teacher_hf_dir


def load_fresh_model(hf_name, teacher_path, device=None):
    model = AutoModelForSeq2SeqLM.from_pretrained(
        hf_name, torch_dtype=torch.bfloat16,
        device_map="cuda" if torch.cuda.is_available() else "auto",
    )
    if teacher_path and os.path.exists(teacher_path):
        model.load_state_dict(torch.load(teacher_path, map_location="cpu"))
        print(f"  Loaded fine-tuned weights from {teacher_path}")
    else:
        print("  No fine-tuned checkpoint - pruning raw pretrained weights")
    if device is not None:
        model = model.to(device)
    return model


def _run_recovery(model, recovery_loader, device, grad_accum, epochs, lr_rec=1e-4):
    print(f"  Recovery (full FT): {epochs} epoch(s)  LR={lr_rec}")
    optimizer = torch.optim.AdamW(model.parameters(), lr=lr_rec)

    for ep in range(epochs):
        model.train()
        if hasattr(model, "gradient_checkpointing_enable"):
            model.gradient_checkpointing_enable()
        optimizer.zero_grad()
        for i, batch in enumerate(recovery_loader):
            ids  = batch["input_ids"].to(device)
            mask = batch["attention_mask"].to(device)
            lbls = batch["labels"].to(device)
            with torch.amp.autocast(device_type="cuda", dtype=torch.bfloat16):
                loss = model(input_ids=ids, attention_mask=mask, labels=lbls).loss / grad_accum
            loss.backward()
            if (i + 1) % grad_accum == 0:
                torch.nn.utils.clip_grad_norm_(model.parameters(), CLIP_NORM)
                optimizer.step()
                optimizer.zero_grad()
            if i % 25 == 0:
                mem = torch.cuda.memory_allocated() / 1024**3 if torch.cuda.is_available() else 0
                print(f"  [recovery ep {ep+1}] batch {i:4d} loss {loss.item()*grad_accum:.4f} GPU {mem:.2f}GB")
    return model


# ════════════════════════════════════════════════════════════════════════════
# METHOD 1: Magnitude pruning
# ════════════════════════════════════════════════════════════════════════════
def apply_magnitude_pruning(model, prune_ratio=PRUNE_RATIO):
    print(f"  [Magnitude] sparsity={prune_ratio:.0%}")
    with torch.no_grad():
        for name, param in get_prunable_params(model):
            flat = param.data.abs().view(-1)
            k = int(prune_ratio * flat.numel())
            if k == 0:
                continue
            threshold = torch.kthvalue(flat, k).values
            param.data.mul_(param.data.abs() > threshold)
    return model


# ════════════════════════════════════════════════════════════════════════════
# METHOD 2: Movement pruning
# ════════════════════════════════════════════════════════════════════════════
def apply_movement_pruning(model, calib_loader, device, prune_ratio=PRUNE_RATIO):
    print(f"  [Movement] collecting gradients (gradient checkpointing enabled) ...")
    model.train()
    if hasattr(model, "gradient_checkpointing_enable"):
        model.gradient_checkpointing_enable()
    model.zero_grad()

    for batch in calib_loader:
        ids  = batch["input_ids"].to(device)
        mask = batch["attention_mask"].to(device)
        lbls = batch["labels"].to(device)
        with torch.amp.autocast(device_type="cuda", dtype=torch.bfloat16):
            loss = model(input_ids=ids, attention_mask=mask, labels=lbls).loss
        loss.backward()
        clear_gpu()

    if hasattr(model, "gradient_checkpointing_disable"):
        model.gradient_checkpointing_disable()

    print(f"  [Movement] applying mask at sparsity={prune_ratio:.0%}")
    with torch.no_grad():
        for name, param in get_prunable_params(model):
            if param.grad is None:
                continue
            scores = (param.data * param.grad).abs()
            k = int(prune_ratio * scores.numel())
            if k == 0:
                param.grad = None
                continue
            threshold = torch.kthvalue(scores.view(-1), k).values
            param.data.mul_(scores > threshold)
            param.grad = None

    model.zero_grad()
    clear_gpu()
    return model


# ════════════════════════════════════════════════════════════════════════════
# METHOD 3: SparseGPT
# ════════════════════════════════════════════════════════════════════════════
class _SparseGPTLayer:
    def __init__(self, layer: nn.Linear):
        self.layer     = layer
        self.n_cols    = layer.weight.shape[1]
        self.H         = torch.zeros(self.n_cols, self.n_cols, dtype=torch.float32)
        self.n_samples = 0

    def add_batch(self, inp):
        x = inp.reshape(-1, self.n_cols).float().cpu()
        self.H        += x.T @ x
        self.n_samples += x.size(0)

    def prune(self, sparsity, device, block_size=128):
        if self.n_samples == 0:
            return
        W = self.layer.weight.data.clone().float()
        H = (self.H / self.n_samples).to(device)
        H.diagonal().add_(0.01 * H.diagonal().mean())

        try:
            H_inv = torch.cholesky_inverse(torch.linalg.cholesky(H))
        except torch.linalg.LinAlgError:
            H_inv = torch.diag(1.0 / H.diagonal().clamp(min=1e-8))

        mask = torch.zeros_like(W, dtype=torch.bool)
        for start in range(0, self.n_cols, block_size):
            end    = min(start + block_size, self.n_cols)
            W_blk  = W[:, start:end].clone()
            H_blk  = H_inv[start:end, start:end]
            h_diag = H_blk.diagonal().clamp(min=1e-8)

            n_prune = int(sparsity * (end - start))
            if n_prune == 0:
                continue
            scores   = W_blk ** 2 / h_diag.unsqueeze(0)
            thresh   = torch.kthvalue(scores.reshape(-1), n_prune).values
            blk_mask = scores <= thresh
            mask[:, start:end] = blk_mask
            err = (W_blk * blk_mask.float()) / h_diag.unsqueeze(0)
            W[:, start:end] -= err @ H_blk

        W[mask] = 0.0
        self.layer.weight.data = W.to(self.layer.weight.dtype)
        del H, H_inv, W, mask


def apply_sparsegpt(model, calib_loader, device, prune_ratio=PRUNE_RATIO):
    print(f"  [SparseGPT] registering hooks ...")
    sg_layers, hooks = {}, {}

    def make_hook(sg):
        def hook(module, inp, out):
            sg.add_batch(inp[0].detach())
        return hook

    for name, module in model.named_modules():
        if isinstance(module, nn.Linear):
            sg = _SparseGPTLayer(module)
            sg_layers[name] = sg
            hooks[name]     = module.register_forward_hook(make_hook(sg))

    print(f"  [SparseGPT] {len(sg_layers)} Linear layers - calibration pass ...")
    model.eval()
    with torch.no_grad():
        for i, batch in enumerate(calib_loader):
            model(
                input_ids=batch["input_ids"].to(device),
                attention_mask=batch["attention_mask"].to(device),
                labels=batch["labels"].to(device),
            )
            clear_gpu()
            if i % 10 == 0:
                print(f"    calibration batch {i}")

    for h in hooks.values():
        h.remove()

    print(f"  [SparseGPT] pruning at sparsity={prune_ratio:.0%} (layer-by-layer, CPU Hessians) ...")
    for name, sg in sg_layers.items():
        sg.prune(prune_ratio, device)
        clear_gpu()

    return model


# ════════════════════════════════════════════════════════════════════════════
# METHOD 4: HAWQ (bitsandbytes INT8)
# ════════════════════════════════════════════════════════════════════════════
def apply_hawq_bnb(hf_name, teacher_hf_dir=None):
    from transformers import BitsAndBytesConfig
    import bitsandbytes  # noqa

    # Quantize the fine-tuned teacher when we have one exported in HF
    # format; fall back to the raw pretrained model only if unavailable.
    source = teacher_hf_dir if teacher_hf_dir and os.path.isdir(teacher_hf_dir) else hf_name
    print(f"  [bitsandbytes] loading {source} in INT8 ...")
    bnb_config = BitsAndBytesConfig(load_in_8bit=True)
    model = AutoModelForSeq2SeqLM.from_pretrained(
        source, quantization_config=bnb_config, device_map="auto",
    )
    model.eval()
    print(f"  [bitsandbytes] INT8 model ready (fine-tuned weights: {source != hf_name}).")
    return model


# ════════════════════════════════════════════════════════════════════════════
# METHOD 5: ZeroQuant (quanto INT8)
# ════════════════════════════════════════════════════════════════════════════
def apply_zeroquant_quanto(model, device):
    try:
        from quanto import quantize, freeze, qint8
    except ImportError:
        from optimum.quanto import quantize, freeze, qint8

    print(f"  [quanto] applying INT8 weight quantization ...")
    quantize(model, weights=qint8)
    freeze(model)
    model = model.to(device)
    print(f"  [quanto] quantization complete.")
    return model


# ════════════════════════════════════════════════════════════════════════════
# METHOD 6: GAMC -- Teacher(reused) -> Pruned -> Distilled -> Quantized
# ════════════════════════════════════════════════════════════════════════════
def apply_gamc(hf_name, teacher_path, recovery_loader, device, grad_accum, recovery_epochs):
    """
    Stage 1 (Teacher) is already handled by get_teacher_checkpoint, same as
    every other method in this harness. This function runs stages 2-4:
      2. task_weighted_prune + recovery fine-tune (from flanutils)
      3. distillation against the teacher checkpoint (contrastive_distillation_loss)
      4. dynamic INT8 quantization on CPU

    Returns (model, eval_device) -- eval_device is "cpu" since dynamic
    quantization only runs on CPU.
    """
    # ---- Stage 2: task-weighted pruning + recovery fine-tune ----
    print("  [GAMC] Stage 2: task-weighted pruning")
    student = load_fresh_model(hf_name, teacher_path, device=device)
    student = task_weighted_prune(student, recovery_loader, PRUNE_RATIO, device)
    clear_gpu()
    student = _run_recovery(student, recovery_loader, device, grad_accum, epochs=recovery_epochs)
    clear_gpu()

    # Checkpoint the pruned+recovered student before Stage 3 (distillation).
    # Stage 3 briefly holds a second full teacher on GPU alongside the student
    # -- the same memory-pressure pattern behind the earlier silent Stage 3
    # crash (~24hrs lost, no mid-stage checkpoint). This avoids repeating that.
    artifacts_dir = os.path.dirname(teacher_path) if teacher_path else "."
    stage2_ckpt = os.path.join(artifacts_dir, "gamc_stage2_pruned_recovered.pt")
    torch.save(student.state_dict(), stage2_ckpt)
    print(f"  [GAMC] Stage 2 checkpoint saved -> {stage2_ckpt}")

    # ---- Stage 3: distillation against the original teacher ----
    print("  [GAMC] Stage 3: distillation")
    teacher = AutoModelForSeq2SeqLM.from_pretrained(
        hf_name, torch_dtype=torch.bfloat16,
        device_map="cuda" if torch.cuda.is_available() else "auto",
    )
    if teacher_path and os.path.exists(teacher_path):
        teacher.load_state_dict(torch.load(teacher_path, map_location="cpu"))
    teacher.eval()

    optimizer = torch.optim.AdamW(student.parameters(), lr=DISTILL_LR)
    student.train()
    optimizer.zero_grad()
    for i, batch in enumerate(recovery_loader):
        ids  = batch["input_ids"].to(device)
        mask = batch["attention_mask"].to(device)
        lbls = batch["labels"].to(device)
        with torch.no_grad():
            with torch.amp.autocast(device_type="cuda", dtype=torch.bfloat16):
                teacher_logits = teacher(input_ids=ids, attention_mask=mask, labels=lbls).logits
        with torch.amp.autocast(device_type="cuda", dtype=torch.bfloat16):
            out = student(input_ids=ids, attention_mask=mask, labels=lbls)
            d_loss = contrastive_distillation_loss(out.logits, teacher_logits)
            loss = (d_loss + out.loss) / grad_accum
        loss.backward()
        if (i + 1) % grad_accum == 0:
            torch.nn.utils.clip_grad_norm_(student.parameters(), CLIP_NORM)
            optimizer.step()
            optimizer.zero_grad()
        if i % 25 == 0:
            print(f"  [GAMC distill] batch {i} loss {loss.item()*grad_accum:.4f}")

    del teacher
    clear_gpu()

    # ---- Stage 4: dynamic INT8 quantization (CPU only) ----
    print("  [GAMC] Stage 4: dynamic quantization (CPU)")
    student = student.to("cpu").float()
    quantized = torch.quantization.quantize_dynamic(student, {torch.nn.Linear}, dtype=torch.qint8)
    clear_gpu()
    return quantized, "cpu"


# ════════════════════════════════════════════════════════════════════════════
# Harness
# ════════════════════════════════════════════════════════════════════════════
def measure(model, test_loader, tokenizer, device, model_path=None):
    m = evaluate_seq2seq_yelp(model, test_loader, tokenizer, device, model_path)
    return {
        "Accuracy (%)": round(m["accuracy"] * 100, 2),
        "Accuracy OffByOne (%)": round(m["accuracy_off_by_one"] * 100, 2),
        "Macro-F1(%)":  round(m["f1_score"] * 100, 2),
        "Latency(ms)":  round(m["latency_sec_per_sample"] * 1000, 2),
        "Size (MB)":    round(m["model_size_mb"], 2),
    }


def run_method(
    method_name, model_key, hf_name, teacher_path, teacher_hf_dir,
    train_loader, calib_loader, recovery_loader, test_loader,
    tokenizer, device, cfg
):
    print(f"\n{'='*60}")
    print(f"  Method : {method_name.upper()}")
    print(f"  Model  : {model_key}  Sparsity: {PRUNE_RATIO:.0%}")
    print(f"{'='*60}")

    clear_gpu()
    artifacts_dir = f"artifacts-{model_key}-yelp"
    grad_accum = cfg["grad_accum"]
    eval_device = device

    if method_name == "HAWQ":
        model = apply_hawq_bnb(hf_name, teacher_hf_dir)

    elif method_name == "ZeroQuant":
        model = load_fresh_model(hf_name, teacher_path, device=device)
        model = apply_zeroquant_quanto(model, device)

    elif method_name == "GAMC":
        model, eval_device = apply_gamc(
            hf_name, teacher_path, recovery_loader, device, grad_accum, cfg["recovery_epochs"]
        )

    else:
        model = load_fresh_model(hf_name, teacher_path, device=device)

        if method_name == "Magnitude":
            model = apply_magnitude_pruning(model, PRUNE_RATIO)
        elif method_name == "Movement":
            model = apply_movement_pruning(model, calib_loader, device, PRUNE_RATIO)
        elif method_name == "SparseGPT":
            model = apply_sparsegpt(model, calib_loader, device, PRUNE_RATIO)
        else:
            raise ValueError(f"Unknown method: {method_name}")

        clear_gpu()
        model = _run_recovery(model, recovery_loader, device, grad_accum, epochs=cfg["recovery_epochs"])

    clear_gpu()

    tmp_path = os.path.join(artifacts_dir, f"{method_name.lower()}_model.pt")
    try:
        torch.save(model.state_dict(), tmp_path)
        print(f"  Saved -> {tmp_path}")
    except Exception as e:
        print(f"  Save skipped ({e})")
        tmp_path = None

    print("  Evaluating ...")
    result = measure(model, test_loader, tokenizer, eval_device, tmp_path)
    del model
    clear_gpu()
    return result


def write_sheet4_yelp_format(df, path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet4_yelp"
    row_idx = 1
    for method in METHODS:
        if method not in df["Method"].unique():
            continue
        ws.cell(row=row_idx, column=1,
                value=f"Method: {method}, Dataset: YELP (https://huggingface.co/datasets/Yelp/yelp_review_full)")
        row_idx += 1
        for col, header in enumerate(
            ["Model Name", "Accuracy (%)", "Accuracy OffByOne (%)", "Macro-F1(%)", "Latency(ms)", "Size (MB)"], start=1
        ):
            ws.cell(row=row_idx, column=col, value=header)
        row_idx += 1
        sub = df[df["Method"] == method]
        for _, r in sub.iterrows():
            for col, key in enumerate(
                ["Model Name", "Accuracy (%)", "Accuracy OffByOne (%)", "Macro-F1(%)", "Latency(ms)", "Size (MB)"], start=1
            ):
                ws.cell(row=row_idx, column=col, value=r[key])
            row_idx += 1
        row_idx += 1
    wb.save(path)
    print(f"Wrote sheet4_yelp-formatted results -> {path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--methods", nargs="+", default=METHODS, choices=METHODS)
    parser.add_argument(
        "--models", nargs="+",
        default=list(MODEL_CONFIGS.keys()), choices=list(MODEL_CONFIGS.keys())
    )
    parser.add_argument("--output_csv", default="sheet4_yelp_results.csv")
    parser.add_argument(
        "--test_n", type=int, default=None,
        help="Subsample the test set to this many examples before evaluating. "
             "evaluate_seq2seq_yelp generates one sample at a time, so the full "
             "50k-row Yelp test split can take 40+ minutes PER METHOD PER MODEL. "
             "Use a few hundred for iteration; omit for final real numbers."
    )
    parser.add_argument(
        "--force_retrain_teacher", action="store_true",
        help="Retrain the teacher even if a cached checkpoint exists. Needed "
             "to pick up the new higher finetune_epochs -- otherwise an old "
             "1-epoch checkpoint gets silently reused regardless of what "
             "MODEL_CONFIGS now says."
    )
    args = parser.parse_args()

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"Device: {device}")

    train_raw, test_raw = load_yelp_data()
    train_df, val_df, test_df = prepare_dataset_yelp(train_raw, test_raw)
    if args.test_n is not None and args.test_n < len(test_df):
        print(f"  Subsampling test set: {len(test_df)} -> {args.test_n} examples")
        test_df = test_df.sample(args.test_n, random_state=42).reset_index(drop=True)

    all_rows = []
    for model_key in args.models:
        cfg      = MODEL_CONFIGS[model_key]
        hf_name  = cfg["hf_name"]

        tokenizer = AutoTokenizer.from_pretrained(hf_name)

        calib_df    = train_df.iloc[:cfg["calib"]].reset_index(drop=True)
        recovery_df = train_df.sample(cfg["recovery_n"], random_state=42).reset_index(drop=True)
        base_df     = train_df.sample(cfg["train_n"],    random_state=42).reset_index(drop=True)

        train_loader    = make_loader(base_df,     tokenizer, cfg, cfg["batch_size"], shuffle=True)
        calib_loader    = make_loader(calib_df,    tokenizer, cfg, cfg["batch_size"], shuffle=False)
        recovery_loader = make_loader(recovery_df, tokenizer, cfg, cfg["batch_size"], shuffle=True)
        test_loader     = make_loader(test_df,     tokenizer, cfg, cfg["test_batch"], shuffle=False)

        teacher_path, teacher_hf_dir = get_teacher_checkpoint(
            model_key, hf_name, train_loader, device, cfg["grad_accum"],
            cfg["finetune_epochs"], force=args.force_retrain_teacher
        )

        for method_name in args.methods:
            try:
                result = run_method(
                    method_name, model_key, hf_name, teacher_path, teacher_hf_dir,
                    train_loader, calib_loader, recovery_loader, test_loader,
                    tokenizer, device, cfg
                )
                all_rows.append({
                    "Method": method_name,
                    "Model Name": DISPLAY_NAMES[model_key],
                    **result,
                })
            except Exception as e:
                print(f"[FAIL] {method_name} on {model_key}: {e}")
                traceback.print_exc()
            finally:
                pd.DataFrame(all_rows).to_csv(args.output_csv, index=False)
                print(f"Saved progress -> {args.output_csv}")
                gc.collect()
                clear_gpu()

    df = pd.DataFrame(all_rows)
    df.to_csv(args.output_csv, index=False)
    if not df.empty:
        write_sheet4_yelp_format(df, args.output_csv.replace(".csv", "_sheet4_format.xlsx"))
    print(f"\nAll done.\n{df}")


if __name__ == "__main__":
    main()